from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _safe_sheet_name(name):
    invalid = "[]:*?/\\"
    safe = "".join("_" if char in invalid else char for char in name)
    return (safe or "document")[:31]


def _document_number(output_file):
    stem = Path(output_file).stem
    for suffix in ("_after", "_read_pdf_top"):
        if stem.endswith(suffix):
            return stem[: -len(suffix)]
    return stem


def _extended_header(block_no, index):
    if index == 1:
        return f"Block {block_no} extended description"
    return f"Block {block_no} extended description {index}"


def _build_main_row(data):
    headers = ["Headline text", "Description text"]
    row = {
        "Headline text": data.get("headline", ""),
        "Description text": data.get("description", ""),
    }

    for block_no, block in enumerate(data.get("blocks", []), start=1):
        header_name = f"Block {block_no} header"
        description_name = f"Block {block_no} description"
        headers.extend([header_name, description_name])
        row[header_name] = block.get("header", "")
        row[description_name] = block.get("description", "")

        for idx, content in enumerate(block.get("extended_descriptions", []), start=1):
            column = _extended_header(block_no, idx)
            headers.append(column)
            row[column] = content

    return headers, row


def _style_sheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            lines = str(cell.value).splitlines() or [""]
            max_len = max(max_len, max(len(line) for line in lines))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 14), 80)


def _write_input_details(wb, document_number):
    ws = wb.active
    ws.title = "input_details"
    rows = [
        ("documentNumber", document_number),
        ("Table of Content", ""),
        ("Company logo image URL", ""),
        ("Asset type details", ""),
        ("Sub-headline text", ""),
        ("Resources description", ""),
        ("Search keywords", ""),
        ("Summary", ""),
        ("Source file", ""),
    ]

    yellow = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    for row_no, (key, value) in enumerate(rows, start=1):
        ws.cell(row=row_no, column=1, value=key)
        ws.cell(row=row_no, column=2, value=value)
        ws.cell(row=row_no, column=1).font = Font(bold=True)
        if row_no in (4, 5, 6):
            ws.cell(row=row_no, column=1).fill = yellow
            if row_no == 4:
                ws.cell(row=row_no, column=2).fill = yellow

    _style_sheet(ws)


def _write_main_sheet(wb, sheet_name, data):
    ws = wb.create_sheet(_safe_sheet_name(sheet_name))
    headers, row = _build_main_row(data)

        # Header styling
    green_fill = PatternFill(
        fill_type="solid",
        fgColor="008000"      # Dark Green
    )

    white_font = Font(
        color="FFFFFF",
        bold=True,
        size=12
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    for col_no, header in enumerate(headers, start=1):

        header_cell = ws.cell(
            row=1,
            column=col_no,
            value=header
        )

        header_cell.fill = green_fill
        header_cell.font = white_font
        header_cell.alignment = header_alignment

        ws.cell(
            row=2,
            column=col_no,
            value=row.get(header, "")
        )

    # Make header row taller
    ws.row_dimensions[1].height = 25

    _style_sheet(ws)
    line_count = max((len(str(ws.cell(row=2, column=col).value or "").splitlines()) for col in range(1, ws.max_column + 1)), default=1)
    ws.row_dimensions[2].height = min(max(line_count * 14, 45), 300)


def _write_table_sheet(wb, table):
    ws = wb.create_sheet(_safe_sheet_name(table["name"]))
    rows = table.get("rows") or []
    if not rows:
        ws.cell(row=1, column=1, value=table["name"])
    else:
        for row_no, row in enumerate(rows, start=1):
            for col_no, value in enumerate(row, start=1):
                ws.cell(row=row_no, column=col_no, value=value)

    _style_sheet(ws)


def save_excel(data, output_file):
    output_file = Path(output_file)
    document_number = _document_number(output_file)

    wb = Workbook()
    _write_input_details(wb, document_number)
    _write_main_sheet(wb, document_number, data)

    for table in data.get("tables", []):
        _write_table_sheet(wb, table)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)

    print(f"Saved: {output_file}")
